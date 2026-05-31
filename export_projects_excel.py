"""
export_projects_excel.py
Run: python export_projects_excel.py
Output: projects_client_review.xlsx

Requires: pip install openpyxl
"""

import sys
import os

# ── Allow running from any directory ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from projects_data import PROJECTS
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Variable → Group mapping ───────────────────────────────────────────────────
VARIABLE_GROUPS = {
    "micro_market_stage":          "1. Demand & Liquidity",
    "rental_absorption_velocity":  "1. Demand & Liquidity",
    "resale_velocity":             "1. Demand & Liquidity",
    "days_on_market":              "1. Demand & Liquidity",
    "occupancy_vacancy_rate":      "1. Demand & Liquidity",
    "safety_crime_index":          "2. Neighborhood Livability",
    "healthcare_access":           "2. Neighborhood Livability",
    "school_quality":              "2. Neighborhood Livability",
    "air_quality_index":           "2. Neighborhood Livability",
    "beach_coastal_access":        "2. Neighborhood Livability",
    "population_growth_rate":      "3. Demographic & Economic Strength",
    "expat_concentration":         "3. Demographic & Economic Strength",
    "immediate_pipeline_risk":     "4. Supply Pressure & Risk",
    "active_unsold_inventory":     "4. Supply Pressure & Risk",
    "average_delay_duration":      "5. Delivery Track Record",
    "pct_projects_delivered":      "5. Delivery Track Record",
    "completion_consistency":      "5. Delivery Track Record",
    "escrow_quality":              "6. Financial Credibility",
    "debt_cash_position":          "6. Financial Credibility",
    "stalled_projects_count":      "6. Financial Credibility",
    "presales_pct_achieved":       "6. Financial Credibility",
    "construction_quality":        "7. Build Integrity",
    "litigation_history":          "8. Legal & Governance",
    "average_exit_velocity":       "9. Exit Liquidity",
    "flipping_frequency":          "9. Exit Liquidity",
    "comp_capital_appreciation":   "10. Comp Market Performance",
    "comp_rental_yields":          "10. Comp Market Performance",
    "comp_occupancy_rate":         "10. Comp Market Performance",
    "projected_rental_yield":      "11. Projected Rental Yields",
    "str_concentration":           "12. Speculation Risk",
    "investor_owner_ratio":        "12. Speculation Risk",
    "offplan_secondary_dominance": "12. Speculation Risk",
    "infrastructure_proximity":    "13. Infra & Connectivity",
}

VARIABLE_LABELS = {
    "micro_market_stage":          "Micro Market Stage",
    "rental_absorption_velocity":  "Rental Absorption Velocity",
    "resale_velocity":             "Resale Velocity",
    "days_on_market":              "Days on Market",
    "occupancy_vacancy_rate":      "Occupancy / Vacancy Rate",
    "safety_crime_index":          "Safety / Crime Index",
    "healthcare_access":           "Healthcare Access",
    "school_quality":              "School Quality",
    "air_quality_index":           "Air Quality Index",
    "beach_coastal_access":        "Beach / Coastal Access",
    "population_growth_rate":      "Population Growth Rate",
    "expat_concentration":         "Expat Concentration",
    "immediate_pipeline_risk":     "Immediate Pipeline Risk",
    "active_unsold_inventory":     "Active Unsold Inventory",
    "average_delay_duration":      "Average Delay Duration",
    "pct_projects_delivered":      "% Projects Delivered",
    "completion_consistency":      "Completion Consistency",
    "escrow_quality":              "Escrow Quality",
    "debt_cash_position":          "Debt / Cash Position",
    "stalled_projects_count":      "Stalled Projects Count",
    "presales_pct_achieved":       "Presales % Achieved",
    "construction_quality":        "Construction Quality",
    "litigation_history":          "Litigation History",
    "average_exit_velocity":       "Average Exit Velocity",
    "flipping_frequency":          "Flipping Frequency",
    "comp_capital_appreciation":   "Comp Capital Appreciation",
    "comp_rental_yields":          "Comp Rental Yields",
    "comp_occupancy_rate":         "Comp Occupancy Rate",
    "projected_rental_yield":      "Projected Rental Yield",
    "str_concentration":           "STR Concentration",
    "investor_owner_ratio":        "Investor / Owner Ratio",
    "offplan_secondary_dominance": "Off-Plan Secondary Dominance",
    "infrastructure_proximity":    "Infrastructure Proximity",
}

# One background colour per group (light pastels, alternating)
GROUP_COLORS = {
    "1. Demand & Liquidity":               "DDEEFF",
    "2. Neighborhood Livability":          "EEFFDD",
    "3. Demographic & Economic Strength":  "FFF4DD",
    "4. Supply Pressure & Risk":           "FFE4E4",
    "5. Delivery Track Record":            "E8DDFF",
    "6. Financial Credibility":            "DDFFF4",
    "7. Build Integrity":                  "FFF0DD",
    "8. Legal & Governance":               "FFDDE8",
    "9. Exit Liquidity":                   "DDEEFF",
    "10. Comp Market Performance":         "EEFFDD",
    "11. Projected Rental Yields":         "FFF4DD",
    "12. Speculation Risk":                "FFE4E4",
    "13. Infra & Connectivity":            "E8DDFF",
}

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FILL    = PatternFill("solid", fgColor="2E75B6")   # medium blue
TITLE_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
SUBHDR_FILL   = PatternFill("solid", fgColor="D6E4F7")
SUBHDR_FONT   = Font(name="Calibri", bold=True, color="1F3864", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
NONE_FONT     = Font(name="Calibri", size=10, italic=True, color="999999")
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
WRAP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER      = Alignment(horizontal="center", vertical="center")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_cell(ws, row, col, value, fill=None, font=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:      c.fill      = fill
    if font:      c.font      = font
    if alignment: c.alignment = alignment
    c.border = THIN_BORDER
    return c

def body_cell(ws, row, col, value, fill=None, font=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font or BODY_FONT
    c.alignment = alignment or WRAP_LEFT
    c.border    = THIN_BORDER
    if fill: c.fill = fill
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Project Overview
# ══════════════════════════════════════════════════════════════════════════════
def build_overview_sheet(wb):
    ws = wb.active
    ws.title = "Project Overview"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="Pangaea Advisory — Project Data Review")
    t.fill      = TITLE_FILL
    t.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["#", "Project Name", "Country", "City",
               "Price Range", "Stage", "Type", "Highlight"]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 2, col, h, fill=HEADER_FILL, font=HEADER_FONT,
                    alignment=Alignment(horizontal="center", vertical="center",
                                        wrap_text=True))
    ws.row_dimensions[2].height = 22

    # Data rows
    country_order = ["Greece", "Portugal", "UAE", "Thailand"]
    projects_sorted = sorted(
        PROJECTS.items(),
        key=lambda x: (
            country_order.index(x[1]["country"])
            if x[1]["country"] in country_order else 99,
            x[1]["city"],
            x[1]["project_name"],
        )
    )

    country_fills = {
        "Greece":   PatternFill("solid", fgColor="EAF4FB"),
        "Portugal": PatternFill("solid", fgColor="EAFBEA"),
        "UAE":      PatternFill("solid", fgColor="FBF4EA"),
        "Thailand": PatternFill("solid", fgColor="F4EAFB"),
    }

    for idx, (key, p) in enumerate(projects_sorted, 1):
        row = idx + 2
        fill = country_fills.get(p["country"], None)
        vals = [
            idx,
            p.get("project_name", key),
            p.get("country", ""),
            p.get("city", ""),
            p.get("price_range", ""),
            p.get("project_stage", ""),
            p.get("project_type", ""),
            p.get("highlight", ""),
        ]
        alignments = [CENTER, WRAP_LEFT, CENTER, CENTER,
                      CENTER, CENTER, CENTER, WRAP_LEFT]
        for col, (val, aln) in enumerate(zip(vals, alignments), 1):
            body_cell(ws, row, col, val, fill=fill, alignment=aln)
        ws.row_dimensions[row].height = 40

    # Column widths
    widths = [4, 38, 12, 16, 20, 18, 14, 55]
    for col, w in enumerate(widths, 1):
        set_col_width(ws, col, w)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Scores & Source Notes (main review sheet)
# ══════════════════════════════════════════════════════════════════════════════
def build_scores_sheet(wb, projects_sorted):
    ws = wb.create_sheet("Scores & Source Notes")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
                value="Scores & Source Notes — Client Review Sheet")
    t.fill      = TITLE_FILL
    t.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Column headers
    col_headers = [
        "Project Name", "Country", "City",
        "Group", "Variable", "Score (0–100)", "Source Note / Methodology",
    ]
    for col, h in enumerate(col_headers, 1):
        header_cell(ws, 2, col, h, fill=HEADER_FILL, font=HEADER_FONT,
                    alignment=Alignment(horizontal="center", vertical="center",
                                        wrap_text=True))
    ws.row_dimensions[2].height = 22

    data_row = 3
    score_rows = []   # track which rows have numeric scores for colour scale

    for key, p in projects_sorted:
        project_name = p.get("project_name", key)
        country      = p.get("country", "")
        city         = p.get("city", "")
        raw_vars     = p.get("raw_variables", {})

        # Build ordered variable list, grouped
        grouped = {}
        unrecognised = []
        for var_key, var_val in raw_vars.items():
            group = VARIABLE_GROUPS.get(var_key, "Uncategorised")
            grouped.setdefault(group, []).append((var_key, var_val))

        # Sort groups by their number prefix, then vars within group
        sorted_groups = sorted(grouped.keys())

        # Project header row (spans across all columns)
        ws.merge_cells(
            start_row=data_row, start_column=1,
            end_row=data_row, end_column=7
        )
        ph = ws.cell(
            row=data_row, column=1,
            value=f"  {project_name}  |  {country}  |  {city}"
        )
        ph.fill      = PatternFill("solid", fgColor="2E75B6")
        ph.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ph.alignment = Alignment(horizontal="left", vertical="center")
        ph.border    = THIN_BORDER
        ws.row_dimensions[data_row].height = 20
        data_row += 1

        for group_name in sorted_groups:
            vars_in_group = grouped[group_name]
            group_fill = PatternFill("solid",
                                     fgColor=GROUP_COLORS.get(group_name, "F5F5F5"))

            # Group sub-header
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=7
            )
            gh = ws.cell(row=data_row, column=1, value=f"  {group_name}")
            gh.fill      = SUBHDR_FILL
            gh.font      = SUBHDR_FONT
            gh.alignment = Alignment(horizontal="left", vertical="center")
            gh.border    = THIN_BORDER
            ws.row_dimensions[data_row].height = 18
            data_row += 1

            for var_key, var_val in vars_in_group:
                label = VARIABLE_LABELS.get(var_key, var_key.replace("_", " ").title())

                if var_val is None:
                    score      = None
                    note       = "— Not scored (no data available; engine outputs 50 neutral)"
                    score_font = NONE_FONT
                    score_disp = "N/A"
                elif isinstance(var_val, dict):
                    score      = var_val.get("score")
                    note       = var_val.get("source_note", "")
                    score_font = BODY_FONT
                    score_disp = round(score, 1) if score is not None else "N/A"
                else:
                    score      = None
                    note       = str(var_val)
                    score_font = BODY_FONT
                    score_disp = "N/A"

                row_vals = [
                    project_name, country, city,
                    group_name, label, score_disp, note,
                ]
                alignments = [
                    WRAP_LEFT, CENTER, CENTER,
                    WRAP_LEFT, WRAP_LEFT, WRAP_CENTER, WRAP_LEFT,
                ]
                for col, (val, aln) in enumerate(zip(row_vals, alignments), 1):
                    c = ws.cell(row=data_row, column=col, value=val)
                    c.alignment = aln
                    c.border    = THIN_BORDER
                    c.fill      = group_fill
                    c.font      = score_font if col == 6 else BODY_FONT

                if isinstance(score_disp, (int, float)):
                    score_rows.append(data_row)

                ws.row_dimensions[data_row].height = 72
                data_row += 1

        data_row += 1  # blank row between projects

    # Column widths
    col_widths = [36, 11, 14, 28, 28, 12, 90]
    for col, w in enumerate(col_widths, 1):
        set_col_width(ws, col, w)

    # Colour scale on Score column (F) — red=0, yellow=50, green=100
    if score_rows:
        score_range = f"F3:F{data_row}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="FF4444",
                mid_type="num",   mid_value=50,    mid_color="FFD700",
                end_type="num",   end_value=100,   end_color="00AA44",
            )
        )

    ws.freeze_panes = "A3"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Per-Country tabs
# ══════════════════════════════════════════════════════════════════════════════
def build_country_sheet(wb, country, projects_for_country):
    ws = wb.create_sheet(country)
    ws.sheet_view.showGridLines = False

    # All projects for this country side by side, one column per project
    project_list = list(projects_for_country)
    n_projects   = len(project_list)

    # Title
    last_col = get_column_letter(2 + n_projects)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws.cell(row=1, column=1, value=f"{country} — Score Comparison")
    t.fill      = TITLE_FILL
    t.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Header: Group | Variable | Project1 | Project2 | ...
    ws.cell(row=2, column=1, value="Group").fill = HEADER_FILL
    ws.cell(row=2, column=1).font = HEADER_FONT
    ws.cell(row=2, column=1).alignment = CENTER
    ws.cell(row=2, column=1).border = THIN_BORDER

    ws.cell(row=2, column=2, value="Variable").fill = HEADER_FILL
    ws.cell(row=2, column=2).font = HEADER_FONT
    ws.cell(row=2, column=2).alignment = CENTER
    ws.cell(row=2, column=2).border = THIN_BORDER

    for i, (key, p) in enumerate(project_list):
        col = 3 + i
        c = ws.cell(row=2, column=col,
                    value=p.get("project_name", key))
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border    = THIN_BORDER
    ws.row_dimensions[2].height = 36

    # Build sorted variable list from first project (fill gaps with N/A)
    all_vars = []
    seen = set()
    for key, p in project_list:
        for var_key in p.get("raw_variables", {}):
            if var_key not in seen:
                all_vars.append(var_key)
                seen.add(var_key)

    # Sort by group
    all_vars.sort(key=lambda v: VARIABLE_GROUPS.get(v, "ZZZ"))

    data_row = 3
    last_group = None

    for var_key in all_vars:
        group = VARIABLE_GROUPS.get(var_key, "Uncategorised")
        label = VARIABLE_LABELS.get(var_key,
                                    var_key.replace("_", " ").title())
        group_fill = PatternFill("solid",
                                 fgColor=GROUP_COLORS.get(group, "F5F5F5"))

        # Group divider row
        if group != last_group:
            last_col_idx = 2 + n_projects
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=last_col_idx
            )
            gh = ws.cell(row=data_row, column=1, value=f"  {group}")
            gh.fill      = SUBHDR_FILL
            gh.font      = SUBHDR_FONT
            gh.alignment = Alignment(horizontal="left", vertical="center")
            gh.border    = THIN_BORDER
            ws.row_dimensions[data_row].height = 16
            data_row += 1
            last_group = group

        # Group | Variable
        g_cell = ws.cell(row=data_row, column=1, value=group)
        g_cell.fill = group_fill; g_cell.font = BODY_FONT
        g_cell.alignment = WRAP_LEFT; g_cell.border = THIN_BORDER

        l_cell = ws.cell(row=data_row, column=2, value=label)
        l_cell.fill = group_fill; l_cell.font = Font(name="Calibri", size=10, bold=True)
        l_cell.alignment = WRAP_LEFT; l_cell.border = THIN_BORDER

        for i, (key, p) in enumerate(project_list):
            col = 3 + i
            var_val = p.get("raw_variables", {}).get(var_key)

            if var_val is None:
                disp = "N/A"
                f = NONE_FONT
            elif isinstance(var_val, dict):
                s = var_val.get("score")
                disp = round(s, 1) if s is not None else "N/A"
                f = BODY_FONT
            else:
                disp = "N/A"
                f = NONE_FONT

            sc = ws.cell(row=data_row, column=col, value=disp)
            sc.fill = group_fill; sc.font = f
            sc.alignment = WRAP_CENTER; sc.border = THIN_BORDER

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # Colour scale across all score columns
    if n_projects > 0:
        for i in range(n_projects):
            col_letter = get_column_letter(3 + i)
            ws.conditional_formatting.add(
                f"{col_letter}3:{col_letter}{data_row}",
                ColorScaleRule(
                    start_type="num", start_value=0,   start_color="FF4444",
                    mid_type="num",   mid_value=50,    mid_color="FFD700",
                    end_type="num",   end_value=100,   end_color="00AA44",
                )
            )

    # Column widths
    set_col_width(ws, 1, 28)
    set_col_width(ws, 2, 28)
    for i in range(n_projects):
        set_col_width(ws, 3 + i, 18)

    ws.freeze_panes = "C3"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    country_order = ["Greece", "Portugal", "UAE", "Thailand"]

    projects_sorted = sorted(
        PROJECTS.items(),
        key=lambda x: (
            country_order.index(x[1]["country"])
            if x[1]["country"] in country_order else 99,
            x[1]["city"],
            x[1]["project_name"],
        )
    )

    wb = Workbook()

    print("Building Project Overview sheet...")
    build_overview_sheet(wb)

    print("Building Scores & Source Notes sheet...")
    build_scores_sheet(wb, projects_sorted)

    print("Building per-country comparison sheets...")
    for country in country_order:
        country_projects = [
            (k, v) for k, v in projects_sorted if v.get("country") == country
        ]
        if country_projects:
            build_country_sheet(wb, country, country_projects)
            print(f"  + {country} ({len(country_projects)} projects)")

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "projects_client_review.xlsx"
    )
    wb.save(out_path)
    print(f"\nDone! File saved to:\n  {out_path}")


if __name__ == "__main__":
    main()
